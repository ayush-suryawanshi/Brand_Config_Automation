# Load all required files ...
#.........................................................................
import os
import django
from time import time,sleep
from ast import literal_eval
from dotenv import load_dotenv
from openpyxl import load_workbook
from Brand_Config_Approval_Portal.scripts.functions import *
from pathlib import Path

#.......................................................................
load_dotenv() # load all environment variables ...
BASE_DIR = Path(__file__).resolve().parent

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Brand_Config_Approval_Portal.settings")
django.setup()

#..............................................................................................

from core.models import *
#........................................
while True:
    email = Email.objects.filter(brand='oppo',processed=False,rejected=False).order_by('created_at').first()
    if email:
        attachments = email.attachments.all()
        email.processed = True



    start_time = time()

    key_id = 1  # Rotating and Tracking API key . 

    for attachment in attachments:


        context_data = read_file_data(BASE_DIR/"engine_brain"/"oppo_explainer.txt")
        sparcs_fields = read_file_data(BASE_DIR/"engine_brain"/"oppo_sparcs_fields.txt")
        sample_scheme_chunck = read_file_data(BASE_DIR/"engine_brain"/"oppo_sample_scheme_chunck.txt")


        try:
            file_counter = 1
            if attachment:
                
                brand_file_path = attachment.attachment.path

            else:
                continue

            clean_blank_cells(brand_file_path)
            workbook = load_workbook(brand_file_path,data_only=True)
            sheet_list = workbook.sheetnames

            
            try:

                for sheet_name in sheet_list:
        
                    sheet = workbook[sheet_name]

                    if sheet.sheet_state == 'hidden' or sheet.sheet_state =='veryHidden':
                        continue

                    print(f'The sheet we processing : {sheet}')

                    counter = 0
                    sample_chunck = []
                    for row in sheet.iter_rows(values_only=True):
                        sample_chunck.append(row)
                        counter+=1

                        if counter > 5:
                            break

                    print(f"The sample chunck : {sample_chunck}")
                    data_lables = fetch_data_lables(sample_chunck)
                    
                    # if 'Realme' in file:
                    #     data_lables = sample_chunck[0]
                    print(f"The data lables are : {data_lables}")
                    data_lables = literal_eval(data_lables) # fetching data lables 

                    df = []
                    counter = 0
                    for row in sheet.iter_rows(values_only=True):
                        if sheet.row_dimensions[counter].hidden == False:
                            df.append(row)
                            counter+=1


                    for row in df:
                        try:
                            input_data = jsonified_output(datalables=data_lables,datapoints=row)
                            new_row = Input_Row.objects.create(brand='Oppo',jsonified_row_data = input_data)
                        except Exception as e:
                            print(f"The data jsonification failed {e}")
                            continue
                        try:
                            bank_and_tenure = gemini_response('''
                                        {
                                    "Input": "Analyze the following data : ''' +str(input_data)+'''Use the following data for understanding given datapoints : '''+str(context_data)+'''",
                                    "Output creation method": "The output should be in form of a python list with 2 seperate lists inside it as follows :",
                                    "List 1 description": "This list should consider data about all the EMI tenures avaialble based on datapoints analysis.",
                                    "List 1 note 1": "Make sure to consider Fullswipe as a part of tenure list if fullswipe schemes are available (if non-emi options are avaialble that mean 'fullswipe' only and should be considered in List 1). Make sure to only add fullwipe as a tenure if explicit mentions of 'fullswipe' or 'Non-EMI' is available",
                                    "EMI Tenures Whitelist": ["3", "6", "9", "12", "18", "24", "fullswipe"],
                                    "List 1 note 2": "Values like 3m , 3M, 6m , 9M, 18 Months , 24 Months , CC EMI , Fullswipe Discount etc. shouldn't be part of the list 1 , only Whitelisted values are allowed in list 1 which will be either pure numeric values or 'fullswipe'.",
                                    "List 1 note 3": "The EMI tenure if mentioned in the text should be picked irrespective of the fact is any offer or discount for it is mentioned or not."
                                    "List 2 description": "This list should strictly only consider data about banks mentioned in 'Bank Whitelist'. If no banks are to be mentioned then the list strictly should be kept empty.",
                                    "List 2 note 1" : "Pick all the banks in the list 2 except for the ones which aren't mentioned in the 'Bank Whitelist' , pick a bank even if no offer or cashback is available for banks ."
                                    "List 2 Example": ["IDFC CC", "IDFC DC", "SBI CC", "SBI DC", "J&K CC", "Axis CC"],
                                    "Bank Whitelist": [
                                        "HDFC CC", "HDFC DC",
                                        "ICICI CC", "ICICI DC",
                                        "AXIS CC", "AXIS DC",
                                        "UNIPAY CC", "UNIPAY DC",
                                        "CITIBANK CC", "CITIBANK DC",
                                        "STANDARD CHARTERED CC", "STANDARD CHARTERED DC",
                                        "HSBC CC", "HSBC DC",
                                        "KOTAK CC", "KOTAK DC",
                                        "SBI CC", "SBI DC",
                                        "AMEX CC",
                                        "INDUSIND CC", "INDUSIND DC",
                                        "BAJAJ CC", "BAJAJ DC",
                                        "RBL CC", "RBL DC",
                                        "YES CC", "YES DC",
                                        "Bank of Baroda CC", "Bank of Baroda DC",
                                        "FEDERAL CC", "FEDERAL DC",
                                        "J&K CC", "J&K DC",
                                        "AU CC", "AU DC",
                                        "ONECARD CC",
                                        "IDFC FIRST CC", "IDFC FIRST DC"
                                    ],
                                    "Final output": [
                                        "The final should contain List1 and List2 in a python list namely : [List1,List2]",
                                        "Explain the logic behind picking the banks and EMI tenure values."
                                        "Strictly make sure the output should be a python list only with list 1 and list 2 and no python code be a part of response."
                                    ]
                                    }

                                        '''
                                        
                                        ,key_id=key_id)
                            
                            print(bank_and_tenure)
                        
                            key_id+=1
                            if key_id > 7:
                                key_id = 1

                            bank_and_tenure = str(format_filter(bank_and_tenure,'[',']'))
                            scheme_combinations = literal_eval(bank_and_tenure)
                            
                            if len(scheme_combinations[0]) < 1 and len(scheme_combinations[1]) < 1:
                                'We are skipping due to insufficient tenure'
                                continue

                        except Exception as e:
                            print(e)
                            continue
                        
                        try:
                            if len(scheme_combinations[1]) > 0 and len(scheme_combinations[0]) > 0:
                                for bank in scheme_combinations[1]:
                                    try: 

                                        scheme = gemini_response(f'''
                        
                                                Analyze the following data : {input_data}

                                                - Use the following data for contextual understanding : {context_data}.
                                                - The output should consist of all the fields mentioned as follows in the jsonic format : {sparcs_fields}.
                                                - Use the sample output to see what kind of output is desired : {sample_chunck}.
                                                - Configure the output JSONs for {bank} bank and following tenures {scheme_combinations[0]}.

                                                [Note : If 'Fullswipe' is also a part of the tenures list then create a seperate json where transection type
                                                is 'Fullswipe' and that json should have all the fields accordingly configured.]
                                                
                                                The output should be in the discussed jsonic format and a seperatre json should be created for every 
                                                bank and tenure combination .

                                                (Note : If cashback value is similar to this - "1000 (1-21 Mar25) 3000 (22-31 Mar25) and applicable for existing EMI
                                                tenure or fullswipe option and bank then create 2 seperate scheme jsons. One with promotion date 1-21 Mar2025 and 1000 as cashback value
                                                and another with 22-31 Mar25 as promotion date and 3000 as cashback value. Use the similar logic to 
                                                process fields with similar patterns untill stated otherwise.Make use to include all the applicable 
                                                EMI tenures say 3,6,9,12,18,24 and both NCE and LCE )
                                                
                                                Example : If schemes is to be created where bank is 'SBI CC' and tenures
                                                are [3,6,9,12,fullswipe]  , then the output should have seperate json for the following combination :
                                                (3 months, SBI CC)
                                                (6 months, SBI CC)
                                                (9 months, SBI CC)
                                                (12 months, SBI CC)
                                                (fullswipe, SBI CC)
                                                
                                                Make sure to put all the output JSON's in a single python list.  

                                                Clearly describe the logic for picking following field values ['Low Cost EMI subvention','Cashback Brand Subvention','Cashback Bank Subvention',
                                                'EMI Bank Subvention' , 'EMI Brand Subvention']. Make sure the logic explaination is seperate from the output jsons. 

                                                Strictly make sure the understanding should'nt be a part of the response when creating final jsons.
                                                Strictly make sure the explainations should'nt be a part of the response when creating final jsons.
                                            
                                                                
                                            ''',key_id=key_id)

                                        key_id+=1

                                        if key_id > 7:
                                            key_id = 1


                                        output_jsons = extract_json_dicts(scheme)
                                        if output_jsons:
                                            for single_scheme in output_jsons:
                                                new_scheme = Scheme.objects.create(
                                                            new_scheme = Scheme.objects.create(
                                                            input_row=new_row,
                                                            brand='Oppo',
                                                            category=single_scheme['category'],
                                                            model_name=single_scheme['model_name'],
                                                            model_description=single_scheme['model_description'],
                                                            sku=single_scheme['sku'],
                                                            mrp=single_scheme['mrp'],
                                                            mop=single_scheme['mop'],
                                                            bank_name=single_scheme['bank_name'],
                                                            scheme_type=single_scheme['scheme_type'],
                                                            emi_type=single_scheme['emi_type'],
                                                            emi_tenure=single_scheme['emi_tenure'],
                                                            low_cost_emi_subvention=single_scheme['low_cost_emi_subvention'],
                                                            cashback_type=single_scheme['cashback_type'],
                                                            cashback_amount=single_scheme['cashback_amount'],
                                                            max_cashback_amount=single_scheme['max_cashback_amount'],
                                                            promo_start_date=single_scheme['promo_start_date'],
                                                            promo_end_date=single_scheme['promo_end_date'],
                                                            scheme_id=single_scheme['scheme_id'],
                                                            is_emi=single_scheme['is_emi'],
                                                            is_cashback=single_scheme['is_cashback'],
                                                            is_upfront=single_scheme['is_upfront'],
                                                            emi_brand_subvention=single_scheme['emi_brand_subvention'],
                                                            emi_bank_subvention=single_scheme['emi_bank_subvention'],
                                                            cashback_brand_subvention=single_scheme['cashback_brand_subvention'],
                                                            cashback_bank_subvention=single_scheme['cashback_bank_subvention'],
                                                            additional_terms=single_scheme['additional_terms'],
                                                            is_approved=single_scheme['is_approved'],
                                                            jsonified_data=single_scheme['jsonified_data'],

                                                        )
                                                )
                                
                                    except Exception as e:
                                        print(e)
                                        
                        except Exception as e:
                            print(e)
            except Exception as e:
                print(e)             
        except Exception as e:
            print(e)
    sleep(30)