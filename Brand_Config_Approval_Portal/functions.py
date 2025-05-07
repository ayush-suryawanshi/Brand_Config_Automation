# Package Imports 
#............................................................................
import os
from json import loads
from dotenv import load_dotenv
from google import generativeai as genai
from openpyxl import load_workbook
#............................................................................
load_dotenv()

def gemini_response(prompt:str,key_id:int): #untested
    genai.configure(api_key=os.getenv(f'GEMINI_API_KEY_{key_id}'))
    model = genai.GenerativeModel("gemini-2.0-flash")
    response = model.generate_content(prompt)
    return response.text

#........................................................................................

def read_file_data(filepath:str):

    with open(filepath,mode='r',encoding='utf-8',errors="ignore") as file:
        filedata = file.read()
        file.close

    return filedata

#.........................................................................................

def fetch_data_lables(context_data:str):
    data_lables = gemini_response(prompt=f''' Please analyze the data and return a list of labels : {context_data} that 
                            seem to be the most likely ones for this dataset. The response should strictly be in 
                            the form of a Python list containing the column names or labels.''',key_id=8)
    data_lables = str(data_lables)
    response = ''
    add = False
    for i in data_lables:
        if i == '[':
            add = True

        if add == True:
            response+=i 

        if i == ']':
            break

    return str(response)

#....................................................................................................................

def format_filter(content,start_sign:str,end_sign:str):
        start = 0
        end = 0

        index = 0
        for i in content:
            if i == start_sign:
                start = index
                break
            index+=1

        index = 0
        for i in content:
            if i == end_sign:
                end = index
            index+=1

        if end == len(content)-1:
            return content[start:]
        return content[start:end+1]

#..............................................................................................................
def jsonified_output(datalables,datapoints):
    response = gemini_response(f'''
            Convert the following datapoints : {datapoints} and datalables : {datalables} to a json and return the json output. 
            - If the list has 2 occurence of datalable 'Banks' on followed by 'CC EMI' another followed by 'Full Swipe' , 
            in this situation first occurance of 'Banks' will be termed 'CC EMI Eligible Banks' and second occurance as
            'Full Swipe Eligible Banks'.
            - Strictly make sure no python code or any additional text should be the part of output json. 
            - If no value is avaialble for a field strictly set it to 'Null' , don't assume any value.
            ''',key_id=8)

    response = format_filter(response,'{','}')
    response = loads(response)
    return response

#.........................................................................................................

def clean_blank_cells(file_path, output_path=None):
    
    wb = load_workbook(file_path)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = None

    # Save the workbook
    save_path = output_path if output_path else file_path
    wb.save(save_path)

#.......................................................................................................

import re
import json

def extract_json_dicts(text):
    # Match all curly-brace structures including nested ones
    brace_stack = []
    json_candidates = []
    start_idx = None

    for i, char in enumerate(text):
        if char == '{':
            if not brace_stack:
                start_idx = i
            brace_stack.append('{')
        elif char == '}':
            brace_stack.pop()
            if not brace_stack and start_idx is not None:
                json_chunk = text[start_idx:i+1]
                json_candidates.append(json_chunk)
                start_idx = None

    # Parse only valid JSON strings into Python dicts
    valid_dicts = []
    for j in json_candidates:
        try:
            data = json.loads(j)
            if isinstance(data, dict):
                valid_dicts.append(data)
        except json.JSONDecodeError:
            continue

    return valid_dicts


